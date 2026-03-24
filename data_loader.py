"""
PBR 데이터 및 ETF 가격 데이터 로드 및 전처리
"""

import pandas as pd
import numpy as np
from openpyxl import load_workbook
from datetime import datetime
import os

# ETF-산업 매핑
ETF_INDUSTRY_MAP = {
    "기술 하드웨어": "XLK",
    "소프트웨어": "IGV",
    "정보통신서비스": "VOX",
    "미디어산업": "XLC",
    "유통산업": "XRT",
    "제약생명공학": "IBB",
    "원자재산업": "XLB",
    "보험산업": "IAK",
    "가정용품": "XLP",
    "음식료담배": "XLP",
    "식품": "XLP",
    "소비자서비스": "XLY",
    "내구소비재": "XLY",
    "자동차와부품": "XLY",
    "헬스케어": "XLV",
    "다각화금융": "XLF",
    "은행산업": "KBE",
    "자본재산업": "XLI",
    "상업전문서비스": "XLI",
    "에너지산업": "XLE",
    "유틸리티": "XLU",
    "운송": "IYT",
}

EXPECTED_INDUSTRIES = list(ETF_INDUSTRY_MAP.keys())


def load_pbr_data(filepath: str) -> pd.DataFrame:
    """
    PBR 데이터 로드 및 파싱
    
    Parameters:
    -----------
    filepath : str
        US_IND_PBR.xlsx 파일 경로
    
    Returns:
    --------
    pd.DataFrame
        Index: date, Columns: 산업명(한글), Values: PBR
    """
    wb = load_workbook(filepath, read_only=True)
    ws = wb["값"]
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    
    # 헤더 추출 (row index 3, 4 = 4번째, 5번째 줄)
    industry_names = [v for v in rows[3][3:] if v is not None]
    
    # 데이터 추출 (row index 6부터)
    data_rows = []
    for r in rows[6:]:
        if r[2] is not None and isinstance(r[2], datetime):
            data_rows.append([r[2].date()] + [r[i] if i < len(r) else None for i in range(3, 3 + len(industry_names))])
    
    pbr_df = pd.DataFrame(data_rows, columns=["date"] + industry_names)
    pbr_df["date"] = pd.to_datetime(pbr_df["date"])
    pbr_df = pbr_df.set_index("date").sort_index()
    pbr_df = pbr_df.apply(pd.to_numeric, errors="coerce")
    
    return pbr_df


def load_etf_data(filepath: str) -> pd.DataFrame:
    """
    ETF 가격 데이터 로드 (sector_etf_prices.csv - pivot table 형식)
    
    Parameters:
    -----------
    filepath : str
        sector_etf_prices.csv 파일 경로
    
    Returns:
    --------
    pd.DataFrame
        Index: date, Columns: ETF 티커, Values: 가격
    """
    df = pd.read_csv(filepath)
    df["date"] = pd.to_datetime(df["date"])
    df = df.set_index("date").sort_index()
    return df


def align_and_resample(pbr_df: pd.DataFrame, etf_df: pd.DataFrame) -> tuple:
    """
    PBR과 ETF 데이터를 월말 기준으로 리샘플링 및 정렬
    
    Returns:
    --------
    tuple: (pbr_monthly, etf_monthly)
    """
    pbr_monthly = pbr_df.resample("ME").last()
    etf_monthly = etf_df.resample("ME").last()
    
    # 공통 날짜로 align
    common_dates = pbr_monthly.index.intersection(etf_monthly.index)
    pbr_monthly = pbr_monthly.loc[common_dates]
    etf_monthly = etf_monthly.loc[common_dates]
    
    return pbr_monthly, etf_monthly


def create_industry_momentum_map(etf_df: pd.DataFrame) -> pd.DataFrame:
    """
    ETF 모멘텀을 산업별로 매핑 (여러 마기업이 동일 ETF 공유)
    
    Parameters:
    -----------
    etf_df : pd.DataFrame
        월말 기준 ETF 가격 (Index: date, Columns: ETF 티커)
    
    Returns:
    --------
    pd.DataFrame
        Index: date, Columns: 산업명, Values: 모멘텈 (또는 ETF 가격)
    """
    industry_etf = pd.DataFrame(index=etf_df.index)
    
    for industry, etf_ticker in ETF_INDUSTRY_MAP.items():
        if etf_ticker in etf_df.columns:
            industry_etf[industry] = etf_df[etf_ticker]
        else:
            industry_etf[industry] = np.nan
    
    return industry_etf
